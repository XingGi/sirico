# backend/app/routes/risk_management_levels.py
import os
import decimal
import openpyxl
from sqlalchemy import func
from flask import request, jsonify, Blueprint, send_file, current_app
from app.models import (
    db, BasicAssessment, OrganizationalContext, 
    BasicRiskIdentification, BasicRiskAnalysis, RiskMapTemplate, RiskMapLikelihoodLabel, 
    RiskMapImpactLabel, RiskMapLevelDefinition, RiskMapScore, MadyaAssessment, OrganizationalStructureEntry, SasaranOrganisasiKPI, RiskInputMadya, basic_assessment_contexts
)
from flask_jwt_extended import jwt_required, get_jwt_identity
from datetime import datetime
from io import BytesIO
from werkzeug.utils import secure_filename
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from collections import defaultdict

# Membuat Blueprint untuk Risk Management Levels
risk_management_levels_bp = Blueprint('risk_management_levels_bp', __name__)

@risk_management_levels_bp.route('/basic-assessments', methods=['POST'])
@jwt_required()
def create_basic_assessment():
    """Endpoint untuk membuat Asesmen Dasar baru beserta konteks dan risikonya."""
    current_user_id = get_jwt_identity()
    data = request.get_json()

    if not data or not data.get('nama_unit_kerja') or not data.get('nama_perusahaan'):
        return jsonify({"msg": "Nama Unit Kerja dan Nama Perusahaan wajib diisi."}), 400

    new_assessment = BasicAssessment(
        nama_unit_kerja=data['nama_unit_kerja'],
        nama_perusahaan=data['nama_perusahaan'],
        user_id=current_user_id
    )
    db.session.add(new_assessment)
    db.session.flush()

    # Proses data konteks jika ada
    contexts_data = data.get('contexts', [])
    for context_item in contexts_data:
        new_context = OrganizationalContext(
            external_context=context_item.get('external'),
            internal_context=context_item.get('internal'),
            user_id=current_user_id
        )
        db.session.add(new_context)
        # Flush untuk mendapatkan ID sebelum di-commit
        db.session.flush()
        new_assessment.contexts.append(new_context)

    # Proses data identifikasi risiko jika ada
    risks_data = data.get('risks', [])
    frontend_to_db_risk_id_map = {} 
    
    for index, risk_item in enumerate(risks_data):
        try:
            tanggal_obj = datetime.strptime(risk_item.get('tanggal_identifikasi'), '%Y-%m-%d').date()
        except (ValueError, TypeError):
            tanggal_obj = datetime.utcnow().date()
            
        new_risk = BasicRiskIdentification(
            kode_risiko=risk_item.get('kode_risiko'),
            kategori_risiko=risk_item.get('kategori_risiko'),
            unit_kerja=risk_item.get('unit_kerja'),
            sasaran=risk_item.get('sasaran'),
            tanggal_identifikasi=tanggal_obj,
            deskripsi_risiko=risk_item.get('deskripsi_risiko'),
            akar_penyebab=risk_item.get('akar_penyebab'),
            indikator_risiko=risk_item.get('indikator_risiko'),
            internal_control=risk_item.get('internal_control'),
            deskripsi_dampak=risk_item.get('deskripsi_dampak'),
            assessment_id=new_assessment.id
        )
        db.session.add(new_risk)
        db.session.flush()
        frontend_to_db_risk_id_map[index] = new_risk.id

    analyses_data = data.get('analyses', [])
    for analysis_item in analyses_data:
        frontend_risk_id = analysis_item.get('risk_identification_id')
        db_risk_id = frontend_to_db_risk_id_map.get(frontend_risk_id)

        if db_risk_id:
            new_analysis = BasicRiskAnalysis(
                risk_identification_id=db_risk_id,
                probabilitas=analysis_item.get('probabilitas'),
                dampak=analysis_item.get('dampak'),
                probabilitas_kualitatif=analysis_item.get('probabilitas_kualitatif'),
                dampak_finansial=analysis_item.get('dampak_finansial'),
                assessment_id=new_assessment.id
            )
            db.session.add(new_analysis)

    db.session.commit()

    return jsonify({"msg": "Asesmen Dasar berhasil dibuat.", "id": new_assessment.id}), 201


@risk_management_levels_bp.route('/basic-assessments', methods=['GET'])
@jwt_required()
def get_all_basic_assessments():
    """Mengambil semua Asesmen Dasar milik pengguna."""
    current_user_id = get_jwt_identity()
    assessments = BasicAssessment.query.filter_by(user_id=current_user_id).order_by(BasicAssessment.created_at.desc()).all()
    
    assessment_list = [{
        "id": a.id,
        "nama_unit_kerja": a.nama_unit_kerja,
        "nama_perusahaan": a.nama_perusahaan,
        "created_at": a.created_at.isoformat()
    } for a in assessments]
        
    return jsonify(assessment_list)

@risk_management_levels_bp.route('/basic-assessments/<int:assessment_id>', methods=['GET'])
@jwt_required()
def get_basic_assessment_detail(assessment_id):
    """Mengambil detail lengkap dari satu Asesmen Dasar."""
    current_user_id = get_jwt_identity()
    assessment = BasicAssessment.query.filter_by(id=assessment_id, user_id=current_user_id).first_or_404()

    risks_list = [{
        "id": r.id, # Sertakan ID asli dari database
        "kode_risiko": r.kode_risiko,
        "kategori_risiko": r.kategori_risiko,
        "unit_kerja": r.unit_kerja,
        "sasaran": r.sasaran,
        "tanggal_identifikasi": r.tanggal_identifikasi.isoformat(),
        "deskripsi_risiko": r.deskripsi_risiko,
        "akar_penyebab": r.akar_penyebab,
        "indikator_risiko": r.indikator_risiko,
        "internal_control": r.internal_control,
        "deskripsi_dampak": r.deskripsi_dampak
    } for r in assessment.risks]

    # 2. Buat pemetaan dari ID risiko ke index-nya di dalam list
    risk_id_to_index_map = {risk['id']: index for index, risk in enumerate(risks_list)}
    
    return jsonify({
        "id": assessment.id,
        "nama_unit_kerja": assessment.nama_unit_kerja,
        "nama_perusahaan": assessment.nama_perusahaan,
        "contexts": [{
            "external": ctx.external_context,
            "internal": ctx.internal_context
        } for ctx in assessment.contexts],
        "risks": risks_list,
        "analyses": [{
            "risk_identification_id": risk_id_to_index_map.get(a.risk_identification_id),
            "probabilitas": a.probabilitas,
            "dampak": a.dampak,
            "probabilitas_kualitatif": a.probabilitas_kualitatif,
            "dampak_finansial": a.dampak_finansial
        } for a in assessment.risk_analyses if a.risk_identification_id in risk_id_to_index_map]
    }), 200
    
@risk_management_levels_bp.route('/basic-assessments/<int:assessment_id>', methods=['PUT'])
@jwt_required()
def update_basic_assessment(assessment_id):
    """Memperbarui Asesmen Dasar yang ada."""
    current_user_id = get_jwt_identity()
    assessment = BasicAssessment.query.filter_by(id=assessment_id, user_id=current_user_id).first_or_404()
    data = request.get_json()

    assessment.nama_unit_kerja = data.get('nama_unit_kerja', assessment.nama_unit_kerja)
    assessment.nama_perusahaan = data.get('nama_perusahaan', assessment.nama_perusahaan)

    # Strategi "Delete-and-Recreate" untuk data terkait
    assessment.contexts.clear()
    BasicRiskAnalysis.query.filter_by(assessment_id=assessment.id).delete()
    BasicRiskIdentification.query.filter_by(assessment_id=assessment.id).delete()
    db.session.flush()

    # Tambahkan kembali data yang diperbarui (Logika ini sama persis dengan POST)
    contexts_data = data.get('contexts', [])
    for context_item in contexts_data:
        new_context = OrganizationalContext(
            external_context=context_item.get('external'),
            internal_context=context_item.get('internal'),
            user_id=current_user_id
        )
        db.session.add(new_context)
        db.session.flush()
        assessment.contexts.append(new_context)
    
    risks_data = data.get('risks', [])
    frontend_to_db_risk_id_map = {}
    for index, risk_item in enumerate(risks_data):
        try:
            tanggal_obj = datetime.strptime(risk_item.get('tanggal_identifikasi'), '%Y-%m-%d').date()
        except (ValueError, TypeError):
            tanggal_obj = datetime.utcnow().date()
            
        new_risk = BasicRiskIdentification(
            assessment_id=assessment.id,
            kode_risiko=risk_item.get('kode_risiko'),
            kategori_risiko=risk_item.get('kategori_risiko'),
            unit_kerja=risk_item.get('unit_kerja'),
            sasaran=risk_item.get('sasaran'),
            tanggal_identifikasi=tanggal_obj,
            deskripsi_risiko=risk_item.get('deskripsi_risiko'),
            akar_penyebab=risk_item.get('akar_penyebab'),
            indikator_risiko=risk_item.get('indikator_risiko'),
            internal_control=risk_item.get('internal_control'),
            deskripsi_dampak=risk_item.get('deskripsi_dampak')
        )
        db.session.add(new_risk)
        db.session.flush()
        frontend_to_db_risk_id_map[index] = new_risk.id

    analyses_data = data.get('analyses', [])
    for analysis_item in analyses_data:
        frontend_risk_id_index = analysis_item.get('risk_identification_id')
        db_risk_id = frontend_to_db_risk_id_map.get(frontend_risk_id_index)

        if db_risk_id:
            new_analysis = BasicRiskAnalysis(
                assessment_id=assessment.id,
                risk_identification_id=db_risk_id,
                probabilitas=analysis_item.get('probabilitas'),
                dampak=analysis_item.get('dampak'),
                probabilitas_kualitatif=analysis_item.get('probabilitas_kualitatif'),
                dampak_finansial=analysis_item.get('dampak_finansial')
            )
            db.session.add(new_analysis)

    db.session.commit()
    return jsonify({"msg": "Asesmen Dasar berhasil diperbarui."}), 200

@risk_management_levels_bp.route('/basic-assessments/<int:assessment_id>', methods=['DELETE'])
@jwt_required()
def delete_basic_assessment(assessment_id):
    """Menghapus Asesmen Dasar beserta data terkaitnya."""
    current_user_id = get_jwt_identity()
    assessment = BasicAssessment.query.filter_by(id=assessment_id, user_id=current_user_id).first_or_404()

    try:
        # Jika relasi 'contexts' di model BasicAssessment menggunakan cascade="all, delete",
        # menghapus assessment akan otomatis menangani tabel asosiasi.
        # Jika tidak, uncomment baris di bawah ini untuk menghapus relasi many-to-many secara manual:
        # assessment.contexts.clear() # Hapus asosiasi ke OrganizationalContext

        # Hapus assessment utama. Relasi one-to-many (risks, risk_analyses)
        # akan terhapus otomatis jika cascade="all, delete-orphan" ada di model.
        db.session.delete(assessment)
        db.session.commit()
        return jsonify({"msg": "Asesmen Dasar berhasil dihapus."}), 200
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting basic assessment {assessment_id}: {e}") # Log error di server
        return jsonify({"msg": "Gagal menghapus asesmen. Terjadi kesalahan internal."}), 500

@risk_management_levels_bp.route('/basic-assessments/<int:assessment_id>/export', methods=['GET'])
@jwt_required()
def export_basic_assessment_to_excel(assessment_id):
    """Membuat dan mengirim file Excel dari data Asesmen Dasar dengan styling lengkap."""
    current_user_id = get_jwt_identity()
    assessment = BasicAssessment.query.filter_by(id=assessment_id, user_id=current_user_id).first_or_404()

    wb = openpyxl.Workbook()
    
    # Hapus sheet default yang dibuat otomatis
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # --- DEFINISI GAYA (STYLES) ---
    bold_font = Font(bold=True)
    white_bold_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal='center', vertical='center')
    
    teal_fill = PatternFill(start_color="4DB6AC", end_color="4DB6AC", fill_type="solid") # Warna Teal
    light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    dark_red_fill = PatternFill(start_color="A90200", end_color="A90200", fill_type="solid")
    bright_red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    peach_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    header_light_blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    value_light_blue_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- Sheet 1: Konteks Organisasi ---
    ws1 = wb.create_sheet(title="Tugas 1 - Konteks")
    ws1.page_setup.orientation = ws1.ORIENTATION_LANDSCAPE
    
    ws1.append(["Tugas 1 - Penetapan Konteks Organisasi"])
    ws1['A1'].font = bold_font
    ws1.append([])
    ws1.append(["Konteks Eksternal", "Konteks Internal"])
    
    for cell in ws1["3:3"]:
        cell.font = white_bold_font
        cell.fill = teal_fill
        cell.alignment = center_align
        cell.border = thin_border

    for ctx in assessment.contexts:
        ws1.append([ctx.external_context, ctx.internal_context])
    
    for row in ws1.iter_rows(min_row=4, max_row=ws1.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
    
    ws1.column_dimensions['A'].width = 80
    ws1.column_dimensions['B'].width = 80

    # --- Sheet 2: Identifikasi Risiko ---
    ws2 = wb.create_sheet(title="Tugas 2 - Identifikasi")
    ws2.page_setup.orientation = ws2.ORIENTATION_LANDSCAPE
    
    ws2['A1'] = "Tugas 2 - Identifikasi Risiko"
    ws2['A1'].font = bold_font
    ws2.append([]) # Baris kosong
    
    # Header Kompleks (Baris 3 dan 4)
    ws2['A3'] = "Nama Unit Kerja:"
    ws2['A3'].font = bold_font
    ws2['B3'] = assessment.nama_unit_kerja
    ws2['B3'].fill = value_light_blue_fill
    
    ws2['A4'] = "Nama Perusahaan:"
    ws2['A4'].font = bold_font
    ws2['B4'] = assessment.nama_perusahaan
    ws2['B4'].fill = value_light_blue_fill

    # Menggabungkan sel untuk "IDENTIFIKASI RISIKO"
    ws2.merge_cells('C3:K4')
    merged_cell_c3 = ws2['C3']
    merged_cell_c3.value = 'IDENTIFIKASI RISIKO'
    merged_cell_c3.alignment = center_align
    merged_cell_c3.font = bold_font
    merged_cell_c3.fill = header_light_blue_fill

    ws2.append([]) # Baris kosong
    
    # Header Tabel (Baris 6)
    headers2 = ["Kode Risiko", "No.", "Kategori Risiko", "Unit Kerja / Fungsi", "Sasaran", "Tanggal Identifikasi Risiko", "Deskripsi atau Kejadian Risiko", "Akar Penyebab", "Indikator Risiko", "Faktor Positif / Internal Control Yang Ada Saat Ini", "Deskripsi Dampak"]
    ws2.append(headers2)
    for cell in ws2[6:6]: # Target baris ke-6
        cell.font = white_bold_font
        cell.fill = dark_red_fill
        cell.alignment = Alignment(wrap_text=True, horizontal='center')

    # Data
    for i, risk in enumerate(assessment.risks):
        ws2.append([
            risk.kode_risiko, i + 1, risk.kategori_risiko, risk.unit_kerja, 
            risk.sasaran, risk.tanggal_identifikasi.strftime('%d %B %Y'), 
            risk.deskripsi_risiko, risk.akar_penyebab, risk.indikator_risiko, 
            risk.internal_control, risk.deskripsi_dampak
        ])

    # Atur lebar kolom
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws2.column_dimensions[col_letter].width = 25

    # --- Sheet 3: Analisis Risiko ---
    ws3 = wb.create_sheet(title="Tugas 3 - Analisis")
    ws3.page_setup.orientation = ws3.ORIENTATION_LANDSCAPE

    ws3['A1'] = "Tugas 3 - Analisis Risiko"
    ws3['A1'].font = bold_font

    ws3['A3'] = "Nama Unit Kerja:"
    ws3['A3'].font = bold_font
    ws3['B3'] = assessment.nama_unit_kerja
    ws3['B3'].fill = value_light_blue_fill

    ws3['A4'] = "Nama Perusahaan:"
    ws3['A4'].font = bold_font
    ws3['B4'] = assessment.nama_perusahaan
    ws3['B4'].fill = value_light_blue_fill

    ws3.merge_cells('C3:G3')
    ws3['C3'].value = 'ANALISIS RISIKO'
    ws3['C3'].alignment = center_align
    ws3['C3'].font = bold_font
    ws3['C3'].fill = peach_fill

    ws3.merge_cells('C4:G4')
    ws3['C4'].value = 'RISIKO INHERENT'
    ws3['C4'].alignment = center_align
    ws3['C4'].font = white_bold_font
    ws3['C4'].fill = bright_red_fill
    
    ws3.append([]) # Baris kosong (baris 5)
    
    headers3 = ["Deskripsi atau Kejadian Risiko", "Probabilitas (P)", "Dampak (I)", "Skor Risiko Inherent (W)", "Probabilitas Risiko Inherent Kualitatif (%)", "Dampak Finansial Risiko Inherent (Rp)", "Nilai Bersih Risiko Inherent"]
    ws3.append(headers3) # Menambahkan headers ke baris 6
    for cell in ws3[6:6]: # Menargetkan baris ke-6 untuk styling
        cell.font = white_bold_font
        cell.fill = dark_red_fill
        cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    for analysis in assessment.risk_analyses:
        skor = (analysis.probabilitas or 0) * (analysis.dampak or 0)
        nilai_bersih = (analysis.dampak_finansial or 0) * ((analysis.probabilitas_kualitatif or 0) / 100)
        ws3.append([analysis.risk_identification.deskripsi_risiko, analysis.probabilitas, analysis.dampak, skor, analysis.probabilitas_kualitatif, analysis.dampak_finansial, nilai_bersih])
        last_row = ws3.max_row
        ws3.cell(row=last_row, column=6).number_format = '"Rp"#,##0'
        ws3.cell(row=last_row, column=7).number_format = '"Rp"#,##0'
    
    for col_letter in ['A','B', 'C', 'D', 'E', 'F', 'G']:
        ws3.column_dimensions[col_letter].width = 25

    # Simpan ke memori
    in_memory_fp = BytesIO()
    wb.save(in_memory_fp)
    in_memory_fp.seek(0)

    return send_file(
        in_memory_fp,
        as_attachment=True,
        download_name=f"Asesmen_Dasar_{assessment.nama_unit_kerja}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@risk_management_levels_bp.route('/madya-assessments/<int:assessment_id>/export', methods=['GET'])
@jwt_required()
def export_madya_assessment_to_excel(assessment_id):
    """Membuat dan mengirim file Excel dari data Asesmen Madya."""
    current_user_id = get_jwt_identity()
    assessment = MadyaAssessment.query.filter_by(id=assessment_id, user_id=current_user_id).first_or_404()
    template = assessment.risk_map_template # Ambil template terkait
    
    risk_inputs = assessment.risk_inputs

    wb = openpyxl.Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # --- DEFINISI GAYA (mirip export Dasar) ---
    bold_font = Font(bold=True)
    white_bold_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Biru tua
    subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # Biru muda
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    risk_num_font = Font(bold=True, color="000000", size=9) # Hitam, tebal, kecil
    risk_num_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    score_align = Alignment(horizontal='right', vertical='top')

    # --- Sheet 1: Struktur Organisasi ---
    ws1 = wb.create_sheet(title="Struktur Organisasi")
    ws1.append(["Struktur Organisasi - Asesmen:", assessment.nama_asesmen])
    ws1['A1'].font = bold_font
    ws1.append([])
    headers1 = ["No", "Direktorat", "Divisi", "Unit Kerja", "Struktur Organisasi"]
    ws1.append(headers1)
    for cell in ws1[3:3]:
        cell.font = white_bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i, entry in enumerate(assessment.structure_entries):
        ws1.append([i + 1, entry.direktorat, entry.divisi, entry.unit_kerja, ""])
        
    if assessment.structure_image_filename:
        try:
            img_path = os.path.join(current_app.config['UPLOAD_FOLDER'], assessment.structure_image_filename)
            if os.path.exists(img_path):
                img = Image(img_path)
                # Sesuaikan ukuran gambar (opsional, atur sesuai preferensi)
                img.height = 150 # Contoh tinggi 150 pixel
                img.width = 200  # Contoh lebar 200 pixel
                # Tambahkan gambar ke sel E4 (baris data pertama, kolom ke-5)
                ws1.add_image(img, 'E4')
                # Atur tinggi baris agar gambar muat (sesuaikan jika perlu)
                ws1.row_dimensions[4].height = 120 # Contoh tinggi 120 point
            else:
                print(f"File gambar tidak ditemukan: {img_path}")
        except Exception as e:
            print(f"Gagal menambahkan gambar struktur: {e}")

    # Apply border and alignment to data rows
    for row in ws1.iter_rows(min_row=4, max_col=5):
        for cell in row:
            cell.border = thin_border
            if cell.column != 5:
                 cell.alignment = left_align_wrap if cell.column > 1 else center_align

    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['D'].width = 30
    ws1.column_dimensions['E'].width = 30


    # --- Sheet 2: Sasaran & Risk Appetite ---
    ws2 = wb.create_sheet(title="Sasaran KPI Appetite")
    ws2.append(["Sasaran Organisasi / KPI & Risk Appetite - Asesmen:", assessment.nama_asesmen])
    ws2['A1'].font = bold_font
    ws2.append([])
    headers2 = ["No", "Sasaran Organisasi / KPI", "Target Appetite", "Skor Risiko Inheren", "Skor Risiko Residual"]
    ws2.append(headers2)
    for cell in ws2[3:3]:
        cell.font = white_bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for i, entry in enumerate(assessment.sasaran_kpi_entries):
        ws2.append([
            i + 1,
            entry.sasaran_kpi,
            entry.target_level,
            entry.inherent_risk_score,
            entry.residual_risk_score
        ])

    for row in ws2.iter_rows(min_row=4, max_col=5):
        for cell in row:
            cell.border = thin_border
            if cell.column == 2: cell.alignment = left_align_wrap
            else: cell.alignment = center_align

    ws2.column_dimensions['B'].width = 60
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 20

    # --- Sheet 3: Risk Input ---
    ws3 = wb.create_sheet(title="Risk Input")
    ws3.page_setup.orientation = ws3.ORIENTATION_LANDSCAPE # Make it wide
    ws3.append(["Detail Risk Input - Asesmen:", assessment.nama_asesmen])
    ws3['A1'].font = bold_font
    ws3.append([])
    headers3 = [
        "No", "Kode Risiko", "Status Risiko", "Peluang/ Ancaman", "Kategori Risiko", "Unit/Fungsi Kerja",
        "Sasaran", "Tanggal Identifikasi Risiko", "Deskripsi atau Kejadian Risiko", "Akar Penyebab",
        "Indikator Risiko", "Faktor Positif / Internal Control Yang Ada Saat Ini", "Deskripsi Dampak",
        "Probabilitas (P)", "Dampak (I)", "Skor Risiko Inherent", "Probabilitas Risiko Inherent Kualitatif (%)",
        "Dampak Finansial Risiko Inherent (Rp)", "Nilai Bersih Risiko Inheren",
        "Pemilik Risiko", "Jabatan Pemilik Risiko", "No. HP Pemilik Risiko", "E-mail Pemilik Risiko",
        "Strategi", "Penanganan Risiko (Risk Treatment)", "Biaya Penanganan Risiko (Rp)",
        "Penanganan Yang Telah Dilakukan", "Status Penanganan", "Jadwal Mulai Penanganan",
        "Jadwal Selesai Penanganan", "PIC Penanganan",
        "Probabilitas Risiko Residual (P)", "Dampak Risiko Residual (I)", "Skor Risiko Residual",
        "Probabilitas Risiko Residual Kualitatif (%)", "Dampak Finansial Risiko Residual (Rp)",
        "Nilai Bersih Risiko Residual", "Tanggal Review"
    ]
    ws3.append(headers3)
    header_row = ws3[3:3]
    for cell in header_row:
        cell.font = white_bold_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Buat map sasaran_id ke teks sasaran_kpi untuk lookup
    sasaran_map = {s.id: s.sasaran_kpi for s in assessment.sasaran_kpi_entries}

    for i, entry in enumerate(assessment.risk_inputs):
        sasaran_text = sasaran_map.get(entry.sasaran_id, "-")
        kategori_display = entry.kategori_risiko_lainnya if entry.kategori_risiko == 'Lainnya' else entry.kategori_risiko
        tgl_identifikasi_str = entry.tanggal_identifikasi.strftime('%d %B %Y') if entry.tanggal_identifikasi else ""
        tgl_mulai_str = entry.jadwal_mulai_penanganan.strftime('%d %B %Y') if entry.jadwal_mulai_penanganan else ""
        tgl_selesai_str = entry.jadwal_selesai_penanganan.strftime('%d %B %Y') if entry.jadwal_selesai_penanganan else ""
        tgl_review_str = entry.tanggal_review.strftime('%d %B %Y') if entry.tanggal_review else ""

        # --- Urutan data disesuaikan dengan headers3 ---
        ws3.append([
            i + 1, entry.kode_risiko, entry.status_risiko, entry.peluang_ancaman, kategori_display, entry.unit_kerja, # 1-6
            sasaran_text, tgl_identifikasi_str, entry.deskripsi_risiko, entry.akar_penyebab, # 7-10
            entry.indikator_risiko, entry.internal_control, entry.deskripsi_dampak, # 11-13
            entry.inherent_probabilitas, entry.inherent_dampak, entry.inherent_skor, # 14-16
            entry.inherent_prob_kualitatif, entry.inherent_dampak_finansial, entry.inherent_nilai_bersih, # 17-19
            entry.pemilik_risiko, entry.jabatan_pemilik, entry.kontak_pemilik_hp, entry.kontak_pemilik_email, # 20-23
            entry.strategi, entry.rencana_penanganan, entry.biaya_penanganan, # 24-26
            entry.penanganan_dilakukan, entry.status_penanganan, tgl_mulai_str, # 27-29
            tgl_selesai_str, entry.pic_penanganan, # 30-31
            entry.residual_probabilitas, entry.residual_dampak, entry.residual_skor, # 32-34
            entry.residual_prob_kualitatif,   # Kolom 35 (AI) - Probabilitas Kualitatif (%)
            entry.residual_dampak_finansial,  # Kolom 36 (AJ) - Dampak Finansial (Rp)
            entry.residual_nilai_bersih,      # Kolom 37 (AK) - Nilai Bersih
            tgl_review_str
        ])
        
        last_row = ws3.max_row
        ws3.cell(row=last_row, column=17).number_format = '#,##0.0"%"' # Probabilitas Kualitatif Inherent
        ws3.cell(row=last_row, column=18).number_format = '"Rp"#,##0'   # Dampak Finansial Inherent
        ws3.cell(row=last_row, column=19).number_format = '"Rp"#,##0'   # Nilai Bersih Inherent
        ws3.cell(row=last_row, column=26).number_format = '"Rp"#,##0'   # Biaya Penanganan
        ws3.cell(row=last_row, column=35).number_format = '#,##0.0"%"' # Probabilitas Kualitatif Residual
        ws3.cell(row=last_row, column=36).number_format = '"Rp"#,##0_);[Red]("-Rp"#,##0)' # AJ - Dampak Fin Residual
        ws3.cell(row=last_row, column=37).number_format = '"Rp"#,##0_);[Red]("-Rp"#,##0)' # AK - Nilai Bersih Residual
        
    center_cols_risk = [
        1, 8, # No, Tgl Identifikasi
        14, 15, 16, 17, # P(In), I(In), Skor(In), Prob Kual(In)
        29, 30, # Tgl Mulai/Selesai Penanganan
        32, 33, 34, 35, # P(Res), I(Res), Skor(Res), Prob Kual(Res)
        38 # Tgl Review
    ]
    # Kolom mata uang (rata kanan)
    currency_cols_risk = [18, 19, 26, 36, 37]

    # Apply styles and width to Risk Input sheet
    for row_idx in range(4, ws3.max_row + 1):
        ws3.row_dimensions[row_idx].height = 45 # Beri tinggi agar wrap text cukup
        for col_idx in range(1, len(headers3) + 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx in center_cols_risk:
                cell.alignment = center_align
            elif col_idx in currency_cols_risk:
                 cell.alignment = Alignment(horizontal='right', vertical='top') # Rata kanan untuk uang
            else: # Sisanya rata kiri atas wrap
                cell.alignment = left_align_wrap

    # Sesuaikan lebar kolom (perlu di-fine-tune lagi mungkin)
    column_widths = {
        'A': 5, 'B': 15, 'C': 12, 'D': 15, 'E': 20, 'F': 25, 'G': 30, 'H': 15, 'I': 40, 'J': 30, # 1-10
        'K': 30, 'L': 30, 'M': 30, 'N': 8, 'O': 8, 'P': 8, 'Q': 12, 'R': 18, 'S': 18, 'T': 20, # 11-20
        'U': 20, 'V': 15, 'W': 20, 'X': 20, 'Y': 40, 'Z': 18, 'AA': 30, 'AB': 15, 'AC': 15, # 21-29
        'AD': 15, 'AE': 20, 'AF': 8, 'AG': 8, 'AH': 8, 'AI': 12, 'AJ': 18, 'AK': 18, 'AL': 15 # 30-38
    }
    for col_letter, width in column_widths.items():
         try: ws3.column_dimensions[col_letter].width = width
         except KeyError: print(f"Peringatan: Kolom {col_letter} tidak ditemukan di sheet Risk Input.")
        
    # --- Sheet 4: Peta Risiko ---
    ws4 = wb.create_sheet(title="Peta Risiko")
    ws4.append(["Peta Risiko - Asesmen:", assessment.nama_asesmen])
    ws4['A1'].font = bold_font
    ws4.append(["Template:", template.name if template else "Tidak ada template"])
    ws4.append([]) # Baris kosong

    # Persiapan data dari template
    likelihood_labels = {l.level: l.label for l in (template.likelihood_labels if template else [])}
    impact_labels = {i.level: i.label for i in (template.impact_labels if template else [])}
    scores_map = {(s.likelihood_level, s.impact_level): s.score for s in (template.scores if template else [])}
    levels_map = sorted(
        [
            (lvl.min_score, lvl.max_score, lvl.color_hex.replace("#", "FF")) # Ambil warna tanpa #, tambahkan alpha FF
            for lvl in (template.level_definitions if template else [])
        ],
        key=lambda x: x[0] # Urutkan berdasarkan min_score
    )

    def get_color_for_score(score):
        if score is None: return "FFFFFFFF" # Putih jika skor null
        for min_s, max_s, color in levels_map:
            if min_s <= score <= max_s:
                return color
        return "FFFFFFFF" # Default putih jika di luar rentang

    # Persiapan data posisi risk input
    inherent_positions = defaultdict(list)
    residual_positions = defaultdict(list)
    for i, risk in enumerate(risk_inputs):
        if risk.inherent_probabilitas and risk.inherent_dampak:
            inherent_positions[(risk.inherent_probabilitas, risk.inherent_dampak)].append(str(i + 1)) # Simpan nomor urut (string)
        if risk.residual_probabilitas and risk.residual_dampak:
            residual_positions[(risk.residual_probabilitas, risk.residual_dampak)].append(str(i + 1)) # Simpan nomor urut (string)

    # Fungsi untuk menggambar satu peta
    def draw_risk_map(start_row, title, positions_data):
        ws4.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=6)
        title_cell = ws4.cell(row=start_row, column=2, value=title)
        title_cell.font = bold_font
        title_cell.alignment = center_align
        start_row += 1

        # Header Dampak
        prob_label_cell = ws4.cell(row=start_row, column=2, value="PROBABILITAS") # Tulis di B<start_row>
        prob_label_cell.alignment = Alignment(horizontal='center', vertical='center', text_rotation=90)
        ws4.merge_cells(start_row=start_row, start_column=2, end_row=start_row + 4, end_column=2)
        
        for i in range(1, 6):
            col_idx = i + 2
            ws4.cell(row=start_row, column=col_idx, value=impact_labels.get(i, f"Impact {i}")).alignment = center_align
            ws4.cell(row=start_row + 6, column=col_idx, value=i).alignment = center_align # Angka 1-5 di bawah

        dampak_label_cell = ws4.cell(row=start_row+6, column=3, value="DAMPAK") # Tulis di C
        dampak_label_cell.alignment = center_align
        ws4.merge_cells(start_row=start_row+6, start_column=3, end_row=start_row+6, end_column=7)

        # Matriks
        for p in range(5, 0, -1): # Probabilitas 5 ke 1
            row_idx = start_row + (5 - p) + 1
            cell_prob_row = ws4.cell(row=row_idx, column=1, value=f"{likelihood_labels.get(p, f'L {p}')}\n({p})") # Tulis di Kolom A
            cell_prob_row.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for i in range(1, 6): # Dampak 1 ke 5
                col_idx = i + 2
                cell = ws4.cell(row=row_idx, column=col_idx)
                score = scores_map.get((p, i), None) # Ambil skor dari template map
                if score is None and template: # Fallback P*I jika skor tidak ada di template
                    score = p * i
                cell.value = score
                fill_color = get_color_for_score(score)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.border = thin_border
                cell.alignment = score_align

                # Tambahkan nomor risk input
                risk_numbers = positions_data.get((p, i), [])
                if risk_numbers:
                     # Gabungkan nomor dengan koma, tampilkan di tengah
                     cell.value = f"{score}\n\n({', '.join(risk_numbers)})" # Skor di atas, nomor di bawah
                     cell.alignment = risk_num_align # Tengah
                     cell.font = risk_num_font # Font khusus untuk nomor

                ws4.row_dimensions[row_idx].height = 40 # Atur tinggi baris
                ws4.column_dimensions[get_column_letter(col_idx)].width = 15 # Atur lebar kolom
                
        ws4.column_dimensions['A'].width = 25 # Untuk label probabilitas per baris
        ws4.column_dimensions['B'].width = 5

    # Gambar Peta Inheren
    draw_risk_map(start_row=4, title="PETA RISIKO INHEREN", positions_data=inherent_positions)
    draw_risk_map(start_row=15, title="PETA RISIKO RESIDUAL", positions_data=residual_positions)

    # Simpan ke memori
    in_memory_fp = BytesIO()
    wb.save(in_memory_fp)
    in_memory_fp.seek(0)

    return send_file(
        in_memory_fp,
        as_attachment=True,
        # Gunakan nama asesmen di nama file
        download_name=f"Asesmen_Madya_{assessment.nama_asesmen or assessment.id}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
# Peta Risiko
@risk_management_levels_bp.route('/risk-maps', methods=['POST'])
@jwt_required()
def create_risk_map_template():
    data = request.get_json()
    current_user_id = get_jwt_identity()
    if not data or not data.get('name'):
        return jsonify({"msg": "Nama template wajib diisi."}), 400

    new_template = RiskMapTemplate(name=data['name'], description=data.get('description'), user_id=current_user_id)
    db.session.add(new_template)
    db.session.flush()

    for label_data in data.get('likelihood_labels', []):
        db.session.add(RiskMapLikelihoodLabel(template_id=new_template.id, **label_data))
    for label_data in data.get('impact_labels', []):
        db.session.add(RiskMapImpactLabel(template_id=new_template.id, **label_data))
    for definition_data in data.get('level_definitions', []):
        db.session.add(RiskMapLevelDefinition(template_id=new_template.id, **definition_data))
    
    # --- PERUBAHAN: Simpan data skor kustom ---
    for score_data in data.get('scores', []):
        db.session.add(RiskMapScore(template_id=new_template.id, **score_data))

    db.session.commit()
    return jsonify({"msg": "Template berhasil dibuat.", "id": new_template.id}), 201


@risk_management_levels_bp.route('/risk-maps', methods=['GET'])
@jwt_required()
def get_risk_map_templates():
    """Mengambil daftar template (default + milik pengguna)."""
    current_user_id = get_jwt_identity()
    
    # Ambil template default (user_id is None) ATAU yang dimiliki user
    templates = RiskMapTemplate.query.filter(
        (RiskMapTemplate.user_id == current_user_id) | (RiskMapTemplate.is_default == True)
    ).order_by(RiskMapTemplate.is_default.desc(), RiskMapTemplate.name).all()

    result = [
        {"id": t.id, "name": t.name, "description": t.description, "is_default": t.is_default}
        for t in templates
    ]
    return jsonify(result)


@risk_management_levels_bp.route('/risk-maps/<int:template_id>', methods=['GET'])
@jwt_required()
def get_risk_map_template_detail(template_id):
    """Mengambil detail lengkap satu template."""
    template = RiskMapTemplate.query.get_or_404(template_id)
    # (Tambahkan otorisasi jika perlu di masa depan)
    scores = RiskMapScore.query.filter_by(template_id=template.id).all()

    return jsonify({
        "id": template.id,
        "name": template.name,
        "description": template.description,
        "is_default": template.is_default,
        "likelihood_labels": [{"level": l.level, "label": l.label} for l in template.likelihood_labels],
        "impact_labels": [{"level": i.level, "label": i.label} for i in template.impact_labels],
        "level_definitions": [{
            "level_name": d.level_name, "color_hex": d.color_hex,
            "min_score": d.min_score, "max_score": d.max_score
        } for d in template.level_definitions],
        "scores": [{
            "likelihood_level": s.likelihood_level, "impact_level": s.impact_level, "score": s.score
        } for s in scores]
    })

@risk_management_levels_bp.route('/risk-maps/<int:template_id>', methods=['PUT'])
@jwt_required()
def update_risk_map_template(template_id):
    template = RiskMapTemplate.query.get_or_404(template_id)
    current_user_id = get_jwt_identity()

    # Otorisasi: Pastikan user tidak mengedit template default atau milik user lain
    if template.is_default or str(template.user_id) != current_user_id:
        return jsonify({"msg": "Akses ditolak."}), 403

    data = request.get_json()
    
    # Update data induk
    template.name = data.get('name', template.name)
    template.description = data.get('description', template.description)

    # Hapus dan buat ulang data anak (strategi sederhana dan efektif)
    RiskMapLikelihoodLabel.query.filter_by(template_id=template.id).delete()
    RiskMapImpactLabel.query.filter_by(template_id=template.id).delete()
    RiskMapLevelDefinition.query.filter_by(template_id=template.id).delete()
    RiskMapScore.query.filter_by(template_id=template.id).delete()
    
    # Buat ulang seperti di endpoint POST
    for label_data in data.get('likelihood_labels', []):
        db.session.add(RiskMapLikelihoodLabel(template_id=template.id, **label_data))
    for label_data in data.get('impact_labels', []):
        db.session.add(RiskMapImpactLabel(template_id=template.id, **label_data))
    for definition_data in data.get('level_definitions', []):
        db.session.add(RiskMapLevelDefinition(template_id=template.id, **definition_data))
    for score_data in data.get('scores', []):
        db.session.add(RiskMapScore(template_id=template.id, **score_data))

    db.session.commit()
    return jsonify({"msg": "Template berhasil diperbarui."})

@risk_management_levels_bp.route('/risk-maps/<int:template_id>', methods=['DELETE'])
@jwt_required()
def delete_risk_map_template(template_id):
    template = RiskMapTemplate.query.get_or_404(template_id)
    current_user_id = get_jwt_identity()

    if template.is_default or str(template.user_id) != current_user_id:
        return jsonify({"msg": "Akses ditolak. Anda tidak dapat menghapus template default atau milik pengguna lain."}), 403
    
    if template.madya_assessments: # Cek apakah list asesmen yang menggunakan template ini tidak kosong
        assessment_ids = [a.id for a in template.madya_assessments]
        return jsonify({
            "msg": f"Template tidak dapat dihapus karena sedang digunakan oleh Asesmen Madya ID: {', '.join(map(str, assessment_ids))}"
        }), 400
        
    try:
        db.session.delete(template)
        db.session.commit()
        return jsonify({"msg": "Template berhasil dihapus."}), 200 # OK
    except Exception as e:
        db.session.rollback() # Batalkan jika ada error
        print(f"Error deleting template {template_id}: {e}")
        return jsonify({"msg": "Gagal menghapus template. Terjadi kesalahan internal."}), 500

# Endpoint Assessment Madya
@risk_management_levels_bp.route('/madya-assessments', methods=['POST'])
@jwt_required()
def create_madya_assessment():
    """Membuat record Madya Assessment baru (awalan)."""
    current_user_id = get_jwt_identity()
    data = request.get_json() or {}
    
    assessment_name = data.get('nama_asesmen')
    if not assessment_name:
        return jsonify({"msg": "'nama_asesmen' wajib diisi."}), 400
    
    selected_template_id = data.get('risk_map_template_id')
    if not selected_template_id:
        # Contoh: Coba cari template default (misal yg is_default=True atau ID=1)
        default_template = RiskMapTemplate.query.filter_by(is_default=True).first()
        if default_template:
            selected_template_id = default_template.id
        else:
            # Fallback jika tidak ada default, bisa set ke null atau template pertama
            first_template = RiskMapTemplate.query.order_by(RiskMapTemplate.id).first()
            if first_template:
                selected_template_id = first_template.id
    # Untuk saat ini, kita hanya buat record dasarnya
    new_assessment = MadyaAssessment(
        nama_asesmen=assessment_name,
        user_id=current_user_id,
        risk_map_template_id=selected_template_id
    )
    db.session.add(new_assessment)
    db.session.commit()
    # Kembalikan ID agar frontend bisa lanjut ke step berikutnya
    return jsonify({"id": new_assessment.id, "message": "Asesmen Madya baru berhasil dibuat."}), 201

@risk_management_levels_bp.route('/madya-assessments', methods=['GET'])
@jwt_required()
def get_madya_assessments_list():
    """Mengambil daftar semua Asesmen Madya milik pengguna."""
    current_user_id = get_jwt_identity()

    # Query asesmen milik user saat ini, urutkan berdasarkan tanggal terbaru
    assessments = MadyaAssessment.query.filter_by(user_id=current_user_id)\
                                     .order_by(MadyaAssessment.created_at.desc())\
                                     .all()

    # Format data yang akan dikirim ke frontend
    assessment_list = []
    for assessment in assessments:
        assessment_list.append({
            "id": assessment.id,
            "nama_asesmen": assessment.nama_asesmen,
            "created_at": assessment.created_at.isoformat(),
            # Tambahkan field lain jika perlu untuk ditampilkan di list,
            # contoh: nama template jika sudah join
            # "template_name": assessment.risk_map_template.name if assessment.risk_map_template else "N/A"
        })

    return jsonify(assessment_list), 200

@risk_management_levels_bp.route('/madya-assessments/<int:assessment_id>/structure-entries', methods=['POST'])
@jwt_required()
def add_structure_entry(assessment_id):
    current_user_id = get_jwt_identity()
    assessment = MadyaAssessment.query.filter_by(id=assessment_id, user_id=current_user_id).first_or_404()

    # --- PERUBAHAN: Ambil data dari JSON, bukan form-data ---
    data = request.get_json()
    if not data:
         return jsonify({"msg": "Request body harus JSON."}), 400

    direktorat = data.get('direktorat')
    divisi = data.get('divisi')
    unit_kerja = data.get('unit_kerja')

    new_entry = OrganizationalStructureEntry(
        assessment_id=assessment.id,
        direktorat=direktorat,
        divisi=divisi,
        unit_kerja=unit_kerja
    )
    db.session.add(new_entry)
    db.session.commit()

    # Kembalikan data yang baru dibuat agar bisa ditampilkan di tabel frontend
    return jsonify({
        "msg": "Entri struktur berhasil ditambahkan.", 
        "entry": {
            "id": new_entry.id,
            "direktorat": new_entry.direktorat,
            "divisi": new_entry.divisi,
            "unit_kerja": new_entry.unit_kerja
        }
    }), 201
    
@risk_management_levels_bp.route('/madya-assessments/<int:assessment_id>/structure-image', methods=['POST'])
@jwt_required()
def upload_structure_image(assessment_id):
    current_user_id = get_jwt_identity()
    assessment = MadyaAssessment.query.filter_by(id=assessment_id, user_id=current_user_id).first_or_404()

    if 'struktur_organisasi_image' not in request.files:
        return jsonify({"msg": "Tidak ada file gambar yang dikirim."}), 400

    file = request.files['struktur_organisasi_image']
    if file.filename == '':
        return jsonify({"msg": "Nama file tidak boleh kosong."}), 400
        
    if file:
        MAX_FILE_SIZE_BYTES = 1 * 1024 * 1024 # 1MB
        MAX_FILE_SIZE_MB_FOR_MSG = 1
        if file.content_length > MAX_FILE_SIZE_BYTES:
             return jsonify({"msg": f"Ukuran file melebihi {MAX_FILE_SIZE_MB_FOR_MSG}MB."}), 413

        filename = secure_filename(f"struktur_{assessment_id}_{datetime.now().timestamp()}_{file.filename}")
        upload_folder = current_app.config['UPLOAD_FOLDER']
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)
        
        # Hapus gambar lama jika ada
        if assessment.structure_image_filename:
             try:
                 os.remove(os.path.join(upload_folder, assessment.structure_image_filename))
             except OSError:
                 pass # Abaikan jika file tidak ditemukan

        file.save(os.path.join(upload_folder, filename))
        assessment.structure_image_filename = filename
        db.session.commit()
        
        image_url_relative = f"api/uploads/{filename}"

        return jsonify({
            "msg": "Gambar struktur organisasi berhasil diupload.",
            "filename": filename,
            "image_url": image_url_relative # Konsisten pakai 'image_url'
        }), 200
    
    return jsonify({"msg": "Upload gagal."}), 400

@risk_management_levels_bp.route('/madya-assessments/<int:assessment_id>', methods=['GET'])
@jwt_required()
def get_madya_assessment_detail(assessment_id):
    """Mengambil detail Madya Assessment, termasuk entri struktur."""
    current_user_id = get_jwt_identity()
    assessment = MadyaAssessment.query.filter_by(id=assessment_id, user_id=current_user_id).first_or_404()

    structure_entries = [
        {
            "id": entry.id, 
            "direktorat": entry.direktorat, 
            "divisi": entry.divisi, 
            "unit_kerja": entry.unit_kerja
        }
        for entry in assessment.structure_entries
    ]
    
    # Ambil dan format data Sasaran/KPI
    sasaran_entries = [
        {
            "id": entry.id,
            "assessment_id": entry.assessment_id,
            "sasaran_kpi": entry.sasaran_kpi,
            "target_level": entry.target_level,
            "inherent_risk_score": entry.inherent_risk_score,
            "residual_risk_score": entry.residual_risk_score
        }
        for entry in assessment.sasaran_kpi_entries # Akses relasi langsung
    ]
    
    image_url_relative = f"api/uploads/{assessment.structure_image_filename}" if assessment.structure_image_filename else None

    return jsonify({
        "id": assessment.id,
        "nama_asesmen": assessment.nama_asesmen,
        "created_at": assessment.created_at.isoformat(),
        "user_id": assessment.user_id,
        "risk_map_template_id": assessment.risk_map_template_id,
        "structure_image_filename": assessment.structure_image_filename,
        "structure_image_url": image_url_relative, 
        "structure_entries": structure_entries,
        "sasaran_kpi_entries": sasaran_entries,
        "filter_organisasi": assessment.filter_organisasi,
        "filter_direktorat": assessment.filter_direktorat,
        "filter_divisi": assessment.filter_divisi,
        "filter_departemen": assessment.filter_departemen,
    })
    
@risk_management_levels_bp.route('/madya-assessments/<int:assessment_id>/filters', methods=['PUT'])
@jwt_required()
def update_madya_assessment_filters(assessment_id):
    """Mengupdate nilai filter yang tersimpan untuk Asesmen Madya."""
    current_user_id = get_jwt_identity()
    assessment = MadyaAssessment.query.filter_by(id=assessment_id, user_id=current_user_id).first_or_404("Asesmen Madya tidak ditemukan atau bukan milik Anda.")

    data = request.get_json()
    if not data:
        return jsonify({"msg": "Request body tidak boleh kosong."}), 400

    # Update field filter jika ada di data request
    assessment.filter_organisasi = data.get('filter_organisasi', assessment.filter_organisasi)
    assessment.filter_direktorat = data.get('filter_direktorat', assessment.filter_direktorat)
    assessment.filter_divisi = data.get('filter_divisi', assessment.filter_divisi)
    assessment.filter_departemen = data.get('filter_departemen', assessment.filter_departemen)

    try:
        db.session.commit()
        return jsonify({
            "msg": "Filter berhasil disimpan.",
            # Opsional: kembalikan nilai yang disimpan
            "filters": {
                "organisasi": assessment.filter_organisasi,
                "direktorat": assessment.filter_direktorat,
                "divisi": assessment.filter_divisi,
                "departemen": assessment.filter_departemen,
            }
        }), 200
    except Exception as e:
        db.session.rollback()
        print(f"Error saving filters for assessment {assessment_id}: {e}")
        return jsonify({"msg": "Gagal menyimpan filter. Terjadi kesalahan internal."}), 500
    
@risk_management_levels_bp.route('/madya-assessments/<int:assessment_id>', methods=['DELETE'])
@jwt_required()
def delete_madya_assessment(assessment_id):
    """Menghapus Asesmen Madya beserta data terkaitnya."""
    current_user_id = get_jwt_identity()
    assessment = MadyaAssessment.query.filter_by(id=assessment_id, user_id=current_user_id).first_or_404()

    try:
        # Hapus file gambar struktur organisasi jika ada
        if assessment.structure_image_filename:
            try:
                upload_folder = current_app.config['UPLOAD_FOLDER']
                file_path = os.path.join(upload_folder, assessment.structure_image_filename)
                if os.path.exists(file_path):
                     os.remove(file_path)
                else:
                     print(f"Peringatan: File gambar tidak ditemukan saat mencoba menghapus: {file_path}")
            except OSError as e:
                print(f"Error saat menghapus file fisik gambar: {e.strerror}")

        db.session.delete(assessment)
        db.session.commit()
        return jsonify({"msg": "Asesmen Madya berhasil dihapus."}), 200
    except Exception as e:
        db.session.rollback()
        print(f"Error deleting madya assessment {assessment_id}: {e}") # Log error di server
        return jsonify({"msg": "Gagal menghapus asesmen madya. Terjadi kesalahan internal."}), 500
    
@risk_management_levels_bp.route('/madya-assessments/<int:assessment_id>/structure-image', methods=['DELETE'])
@jwt_required()
def delete_structure_image(assessment_id):
    current_user_id = get_jwt_identity()
    assessment = MadyaAssessment.query.filter_by(id=assessment_id, user_id=current_user_id).first_or_404()

    if not assessment.structure_image_filename:
        return jsonify({"msg": "Tidak ada gambar untuk dihapus."}), 404

    # Hapus file fisik
    try:
        upload_folder = current_app.config['UPLOAD_FOLDER']
        file_path = os.path.join(upload_folder, assessment.structure_image_filename)
        if os.path.exists(file_path):
             os.remove(file_path)
        else:
             print(f"Peringatan: File gambar tidak ditemukan saat mencoba menghapus: {file_path}")
    except OSError as e:
        print(f"Error saat menghapus file fisik: {e.strerror}")
        # Kita lanjutkan saja untuk menghapus referensi di DB

    # Hapus referensi di database
    assessment.structure_image_filename = None
    db.session.commit()

    return jsonify({"msg": "Gambar struktur organisasi berhasil dihapus."}), 200

@risk_management_levels_bp.route('/madya-assessments/<int:assessment_id>/template', methods=['PUT'])
@jwt_required()
def update_madya_assessment_template(assessment_id):
    """Mengupdate template peta risiko yang digunakan untuk Asesmen Madya."""
    current_user_id = get_jwt_identity()
    assessment = MadyaAssessment.query.filter_by(id=assessment_id, user_id=current_user_id).first_or_404("Asesmen Madya tidak ditemukan atau bukan milik Anda.")

    data = request.get_json()
    new_template_id = data.get('risk_map_template_id')

    if new_template_id is None:
        return jsonify({"msg": "'risk_map_template_id' wajib diisi."}), 400

    # Validasi apakah template ID valid (opsional tapi bagus)
    template_exists = RiskMapTemplate.query.get(new_template_id)
    if not template_exists:
        return jsonify({"msg": f"Template dengan ID {new_template_id} tidak ditemukan."}), 404

    assessment.risk_map_template_id = new_template_id
    db.session.commit()

    return jsonify({"msg": "Template peta risiko berhasil diperbarui."})

@risk_management_levels_bp.route('/structure-entries/<int:entry_id>', methods=['PUT', 'DELETE'])
@jwt_required()
def manage_structure_entry(entry_id):
    current_user_id = get_jwt_identity()
    entry = OrganizationalStructureEntry.query.get_or_404(entry_id)

    # Otorisasi: Pastikan entry ini milik asesmen user
    assessment = MadyaAssessment.query.get_or_404(entry.assessment_id)
    if str(assessment.user_id) != current_user_id:
        return jsonify({"msg": "Akses ditolak."}), 403

    if request.method == 'PUT':
        data = request.get_json()
        entry.direktorat = data.get('direktorat', entry.direktorat)
        entry.divisi = data.get('divisi', entry.divisi)
        entry.unit_kerja = data.get('unit_kerja', entry.unit_kerja)
        db.session.commit()
        # Kembalikan data yang sudah diupdate
        return jsonify({
            "msg": "Entri berhasil diperbarui.",
            "entry": { "id": entry.id, "direktorat": entry.direktorat, "divisi": entry.divisi, "unit_kerja": entry.unit_kerja }
        })

    if request.method == 'DELETE':
        db.session.delete(entry)
        db.session.commit()
        return jsonify({"msg": "Entri berhasil dihapus."})
    
@risk_management_levels_bp.route('/madya-assessments/<int:assessment_id>/sasaran-kpi', methods=['POST'])
@jwt_required()
def add_sasaran_kpi(assessment_id):
    """Menambahkan entri Sasaran Organisasi/KPI baru ke Asesmen Madya."""
    current_user_id = get_jwt_identity()
    # Pastikan asesmen ada dan milik user
    assessment = MadyaAssessment.query.filter_by(id=assessment_id, user_id=current_user_id).first_or_404("Asesmen Madya tidak ditemukan atau bukan milik Anda.")

    data = request.get_json()
    sasaran_text = data.get('sasaran_kpi')

    if not sasaran_text:
        return jsonify({"msg": "Teks 'sasaran_kpi' wajib diisi."}), 400

    new_sasaran = SasaranOrganisasiKPI(
        assessment_id=assessment.id,
        sasaran_kpi=sasaran_text
        # target_level, inherent_risk_score, residual_risk_score akan null by default
    )
    db.session.add(new_sasaran)
    db.session.commit()

    # Kembalikan data yang baru dibuat
    return jsonify({
        "msg": "Sasaran Organisasi/KPI berhasil ditambahkan.",
        "entry": {
            "id": new_sasaran.id,
            "assessment_id": new_sasaran.assessment_id,
            "sasaran_kpi": new_sasaran.sasaran_kpi,
            "target_level": new_sasaran.target_level,
            "inherent_risk_score": new_sasaran.inherent_risk_score,
            "residual_risk_score": new_sasaran.residual_risk_score
        }
    }), 201

@risk_management_levels_bp.route('/madya-assessments/<int:assessment_id>/sasaran-kpi', methods=['GET'])
@jwt_required()
def get_sasaran_kpi_list(assessment_id):
    """Mengambil semua entri Sasaran Organisasi/KPI untuk Asesmen Madya."""
    current_user_id = get_jwt_identity()
    assessment = MadyaAssessment.query.filter_by(id=assessment_id, user_id=current_user_id).first_or_404("Asesmen Madya tidak ditemukan atau bukan milik Anda.")

    # Ambil semua entri sasaran terkait, urutkan berdasarkan ID (atau kriteria lain jika perlu)
    sasaran_entries = SasaranOrganisasiKPI.query.filter_by(assessment_id=assessment.id).order_by(SasaranOrganisasiKPI.id).all()

    result = [{
        "id": entry.id,
        "assessment_id": entry.assessment_id,
        "sasaran_kpi": entry.sasaran_kpi,
        "target_level": entry.target_level,
        "inherent_risk_score": entry.inherent_risk_score,
        "residual_risk_score": entry.residual_risk_score
    } for entry in sasaran_entries]

    return jsonify(result)

@risk_management_levels_bp.route('/sasaran-kpi/<int:sasaran_id>/target', methods=['PUT'])
@jwt_required()
def update_sasaran_target_level(sasaran_id):
    """Memperbarui target level (Risk Appetite) untuk satu Sasaran/KPI."""
    current_user_id = get_jwt_identity()
    sasaran_entry = SasaranOrganisasiKPI.query.get_or_404(sasaran_id)

    # Otorisasi: Pastikan entri ini milik asesmen user
    assessment = MadyaAssessment.query.get_or_404(sasaran_entry.assessment_id)
    if str(assessment.user_id) != current_user_id:
        return jsonify({"msg": "Akses ditolak."}), 403

    data = request.get_json()
    new_target = data.get('target_level')

    # Validasi sederhana (opsional, bisa lebih ketat)
    allowed_targets = ['R', 'L', 'M', 'H', 'E', None, '']
    if new_target not in allowed_targets:
        return jsonify({"msg": "Nilai target_level tidak valid."}), 400

    sasaran_entry.target_level = new_target if new_target else None # Simpan None jika string kosong
    db.session.commit()

    # Kembalikan data yang sudah diupdate
    return jsonify({
        "msg": "Target level berhasil diperbarui.",
        "entry": {
            "id": sasaran_entry.id,
            "target_level": sasaran_entry.target_level
            # Sertakan field lain jika perlu di-update di frontend
        }
    })

@risk_management_levels_bp.route('/sasaran-kpi/<int:sasaran_id>', methods=['DELETE'])
@jwt_required()
def delete_sasaran_kpi(sasaran_id):
    """Menghapus satu entri Sasaran Organisasi/KPI."""
    current_user_id = get_jwt_identity()
    sasaran_entry = SasaranOrganisasiKPI.query.get_or_404(sasaran_id)

    # Otorisasi: Pastikan entri ini milik asesmen user yang sedang login
    assessment = MadyaAssessment.query.get_or_404(sasaran_entry.assessment_id)
    if str(assessment.user_id) != current_user_id:
        return jsonify({"msg": "Akses ditolak."}), 403

    db.session.delete(sasaran_entry)
    db.session.commit()
    return jsonify({"msg": "Entri Sasaran Organisasi/KPI berhasil dihapus."})

# Risk Input
def parse_date_or_none(date_str):
    if not date_str:
        return None
    try:
        return datetime.strptime(date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return None # Return None if format is invalid or input is not string

# Helper function to safely convert float or return None
def parse_float_or_none(value):
    if value in [None, '']:
        return None
    try:
        # Hapus pemisah ribuan jika ada (misal dari input frontend)
        cleaned_value = str(value).replace('.', '').replace(',', '.')
        return float(cleaned_value)
    except (ValueError, TypeError):
        return None

# Helper function to safely convert int or return None
def parse_int_or_none(value):
    if value in [None, '']:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None

# Helper function to calculate scores and net values
# def calculate_risk_values(data, template_id):
#     inherent_p = parse_int_or_none(data.get('inherent_probabilitas'))
#     inherent_i = parse_int_or_none(data.get('inherent_dampak'))
#     inherent_prob_kualitatif = parse_float_or_none(data.get('inherent_prob_kualitatif'))
#     inherent_dampak_finansial = parse_float_or_none(data.get('inherent_dampak_finansial'))

#     residual_p = parse_int_or_none(data.get('residual_probabilitas'))
#     residual_i = parse_int_or_none(data.get('residual_dampak'))
#     residual_prob_kualitatif = parse_float_or_none(data.get('residual_prob_kualitatif'))
#     residual_dampak_finansial = parse_float_or_none(data.get('residual_dampak_finansial'))
    
#     calculated = {
#         'inherent_skor': None,
#         'inherent_nilai_bersih': None,
#         'residual_skor': None,
#         'residual_nilai_bersih': None
#     }
    
#     def get_score_from_template(p, i):
#         if p is not None and i is not None and template_id is not None:
#             score_entry = RiskMapScore.query.filter_by(
#                 template_id=template_id,
#                 likelihood_level=p,
#                 impact_level=i
#             ).first()
#             return score_entry.score if score_entry else None
#         elif p is not None and i is not None:
#              print(f"PERINGATAN: Tidak ada template_id ({template_id}) atau skor tidak ditemukan untuk P={p}, I={i}. Menggunakan P*I.")
#              return p * i
#         return None

#     calculated['inherent_skor'] = get_score_from_template(inherent_p, inherent_i)
    
#     if inherent_prob_kualitatif is not None and inherent_dampak_finansial is not None:
#         calculated['inherent_nilai_bersih'] = inherent_dampak_finansial * (inherent_prob_kualitatif / 100.0)

#     calculated['residual_skor'] = get_score_from_template(residual_p, residual_i)

#     if residual_prob_kualitatif is not None and residual_dampak_finansial is not None:
#         dampak_fin_res = residual_dampak_finansial if residual_dampak_finansial is not None else inherent_dampak_finansial
#         if dampak_fin_res is not None:
#              calculated['residual_nilai_bersih'] = dampak_fin_res * (residual_prob_kualitatif / 100.0)

#     return calculated

def calculate_scores_from_template(data, template_id):
    inherent_p = parse_int_or_none(data.get('inherent_probabilitas'))
    inherent_i = parse_int_or_none(data.get('inherent_dampak'))
    residual_p = parse_int_or_none(data.get('residual_probabilitas'))
    residual_i = parse_int_or_none(data.get('residual_dampak'))

    scores = {'inherent_skor': None, 'residual_skor': None}

    def get_score(p, i):
        if p is not None and i is not None and template_id is not None:
            score_entry = RiskMapScore.query.filter_by(
                template_id=template_id, likelihood_level=p, impact_level=i
            ).first()
            if score_entry: return score_entry.score
        # Fallback P*I
        if p is not None and i is not None and 1 <= p <= 5 and 1 <= i <= 5:
             print(f"Fallback score calculation P*I for P={p}, I={i} (Template ID: {template_id})")
             return p * i
        return None

    scores['inherent_skor'] = get_score(inherent_p, inherent_i)
    scores['residual_skor'] = get_score(residual_p, residual_i)
    return scores

# --- Helper function to format entry for JSON response ---
def format_risk_input_entry(entry):
    """Mengubah objek RiskInputMadya menjadi dictionary yang JSON-friendly."""
    if not entry:
        return None
    entry_dict = {c.name: getattr(entry, c.name) for c in entry.__table__.columns}

    for date_field in ['tanggal_identifikasi', 'jadwal_mulai_penanganan', 'jadwal_selesai_penanganan', 'tanggal_review']:
        date_value = getattr(entry, date_field)
        entry_dict[date_field] = date_value.isoformat() if date_value else None

    # Pastikan field angka adalah float/int atau null (termasuk konversi dari Decimal jika ada)
    for num_field in [
        'inherent_probabilitas', 'inherent_dampak', 'inherent_skor',
        'inherent_prob_kualitatif', 'inherent_dampak_finansial', 'inherent_nilai_bersih',
        'biaya_penanganan',
        'residual_probabilitas', 'residual_dampak', 'residual_skor',
        'residual_prob_kualitatif', 'residual_dampak_finansial', 'residual_nilai_bersih',
        'sasaran_id'
    ]:
        value = getattr(entry, num_field)
        if isinstance(value, decimal.Decimal):
             entry_dict[num_field] = float(value)
        elif value is not None:
             try:
                 if '.' in str(value): 
                     entry_dict[num_field] = float(value)
                 else:
                     entry_dict[num_field] = int(value)
             except (ValueError, TypeError):
                 entry_dict[num_field] = None
        else:
            entry_dict[num_field] = None

    return entry_dict

def update_sasaran_kpi_scores(sasaran_id):
    """Menghitung dan memperbarui skor inheren & residual tertinggi untuk Sasaran/KPI."""
    if not sasaran_id:
        return

    sasaran_entry = SasaranOrganisasiKPI.query.get(sasaran_id)
    if not sasaran_entry:
        return

    # Cari skor inheren tertinggi dari semua risk input yang terkait
    max_inherent_score = db.session.query(func.max(RiskInputMadya.inherent_skor))\
        .filter(RiskInputMadya.sasaran_id == sasaran_id)\
        .scalar()

    # Cari skor residual tertinggi dari semua risk input yang terkait
    # Pastikan hanya menghitung jika residual_skor tidak null
    max_residual_score = db.session.query(func.max(RiskInputMadya.residual_skor))\
        .filter(RiskInputMadya.sasaran_id == sasaran_id, RiskInputMadya.residual_skor.isnot(None))\
        .scalar()

    sasaran_entry.inherent_risk_score = max_inherent_score
    # Set residual score ke null jika tidak ada risk input dengan residual score
    sasaran_entry.residual_risk_score = max_residual_score if max_residual_score is not None else None

    # Tidak perlu db.session.add() karena objek sudah ada
    # db.session.commit() akan dipanggil di fungsi utama (add/update risk input)
    print(f"Updated scores for SasaranKPI ID {sasaran_id}: Inherent={max_inherent_score}, Residual={max_residual_score}")

@risk_management_levels_bp.route('/madya-assessments/<int:assessment_id>/risk-inputs', methods=['POST'])
@jwt_required()
def add_risk_input(assessment_id):
    """Menambahkan entri Risk Input baru ke Asesmen Madya."""
    current_user_id = get_jwt_identity()
    assessment = MadyaAssessment.query.filter_by(id=assessment_id, user_id=current_user_id).first_or_404("Asesmen Madya tidak ditemukan atau bukan milik Anda.")
    data = request.get_json()

    if not data or not data.get('deskripsi_risiko') or not data.get('kategori_risiko'):
        return jsonify({"msg": "Deskripsi Risiko dan Kategori Risiko wajib diisi."}), 400

    kategori = data.get('kategori_risiko')
    kategori_lainnya = data.get('kategori_risiko_lainnya') if kategori == 'Lainnya' else None

    # calculated_values = calculate_risk_values(data, assessment.risk_map_template_id)
    
    # 1. Parse semua input numerik yang relevan
    inherent_p = parse_int_or_none(data.get('inherent_probabilitas'))
    inherent_i = parse_int_or_none(data.get('inherent_dampak'))
    inherent_prob_kualitatif = parse_float_or_none(data.get('inherent_prob_kualitatif'))
    inherent_dampak_finansial = parse_float_or_none(data.get('inherent_dampak_finansial'))
    residual_p = parse_int_or_none(data.get('residual_probabilitas'))
    residual_i = parse_int_or_none(data.get('residual_dampak'))
    residual_prob_kualitatif = parse_float_or_none(data.get('residual_prob_kualitatif'))
    biaya_penanganan = parse_float_or_none(data.get('biaya_penanganan'))

    # 2. Hitung Skor menggunakan template
    calculated_scores = calculate_scores_from_template(data, assessment.risk_map_template_id)
    inherent_skor = calculated_scores.get('inherent_skor')
    residual_skor = calculated_scores.get('residual_skor')

    # 3. Hitung Nilai Bersih Inheren
    inherent_nilai_bersih = None
    if inherent_prob_kualitatif is not None and inherent_dampak_finansial is not None:
        inherent_nilai_bersih = inherent_dampak_finansial * (inherent_prob_kualitatif / 100.0)

    # 4. Hitung Dampak Finansial Residual (sesuai logika frontend)
    calculated_residual_dampak_finansial = None
    if residual_prob_kualitatif is not None and inherent_dampak_finansial is not None:
        # Asumsi: Dampak finansial residual dihitung dari inheren * probabilitas residual
        calculated_residual_dampak_finansial = inherent_dampak_finansial * (residual_prob_kualitatif / 100.0)

    # 5. Hitung Nilai Bersih Residual (sesuai logika frontend)
    residual_nilai_bersih = None
    if residual_prob_kualitatif is not None and calculated_residual_dampak_finansial is not None:
        # Nilai bersih residual = Dampak Fin Residual * Prob Kual Residual
        residual_nilai_bersih = calculated_residual_dampak_finansial * (residual_prob_kualitatif / 100.0)
        # ATAU JIKA LOGIKA FRONTEND BERBEDA (misal: Dampak Inheren * Prob Kual Res * Prob Kual Res), sesuaikan di sini.
        # Berdasarkan kode frontend Anda, sepertinya:
        # residual_nilai_bersih = calculated_residual_dampak_finansial * (residual_prob_kualitatif / 100.0)

    new_risk_input = RiskInputMadya(
        assessment_id=assessment_id,
        sasaran_id=parse_int_or_none(data.get('sasaran_id')), # Ambil dari form nanti
        kode_risiko=data.get('kode_risiko'),
        status_risiko=data.get('status_risiko', 'Risiko Aktif'),
        peluang_ancaman=data.get('peluang_ancaman', 'Ancaman'),
        kategori_risiko=kategori,
        kategori_risiko_lainnya=kategori_lainnya,
        unit_kerja=data.get('unit_kerja'),
        tanggal_identifikasi=parse_date_or_none(data.get('tanggal_identifikasi')) or datetime.utcnow().date(),
        deskripsi_risiko=data.get('deskripsi_risiko'),
        akar_penyebab=data.get('akar_penyebab'),
        indikator_risiko=data.get('indikator_risiko'),
        internal_control=data.get('internal_control'),
        deskripsi_dampak=data.get('deskripsi_dampak'),
        pemilik_risiko=data.get('pemilik_risiko'),
        jabatan_pemilik=data.get('jabatan_pemilik'),
        kontak_pemilik_hp=data.get('kontak_pemilik_hp'),
        kontak_pemilik_email=data.get('kontak_pemilik_email'),
        strategi=data.get('strategi'),
        rencana_penanganan=data.get('rencana_penanganan'),
        penanganan_dilakukan=data.get('penanganan_dilakukan'),
        status_penanganan=data.get('status_penanganan'),
        jadwal_mulai_penanganan=parse_date_or_none(data.get('jadwal_mulai_penanganan')),
        jadwal_selesai_penanganan=parse_date_or_none(data.get('jadwal_selesai_penanganan')),
        pic_penanganan=data.get('pic_penanganan'),
        tanggal_review=parse_date_or_none(data.get('tanggal_review')),
        
        inherent_probabilitas=inherent_p,
        inherent_dampak=inherent_i,
        inherent_skor=inherent_skor, # Dari calculate_scores_from_template
        inherent_prob_kualitatif=inherent_prob_kualitatif,
        inherent_dampak_finansial=inherent_dampak_finansial,
        inherent_nilai_bersih=inherent_nilai_bersih, # Dari perhitungan manual
        biaya_penanganan=biaya_penanganan,
        residual_probabilitas=residual_p,
        residual_dampak=residual_i,
        residual_skor=residual_skor, # Dari calculate_scores_from_template
        residual_prob_kualitatif=residual_prob_kualitatif,
        residual_dampak_finansial=calculated_residual_dampak_finansial, # Simpan nilai dari form
        residual_nilai_bersih=residual_nilai_bersih,
    )

    db.session.add(new_risk_input)
    db.session.commit()
    db.session.refresh(new_risk_input)
    
    update_sasaran_kpi_scores(new_risk_input.sasaran_id)
    db.session.commit()

    # Kembalikan data lengkap setelah disimpan
    return jsonify({
        "msg": "Risk Input berhasil ditambahkan.",
        "entry": format_risk_input_entry(new_risk_input)
    }), 201


@risk_management_levels_bp.route('/madya-assessments/<int:assessment_id>/risk-inputs', methods=['GET'])
@jwt_required()
def get_risk_inputs(assessment_id):
    """Mengambil semua entri Risk Input untuk Asesmen Madya."""
    current_user_id = get_jwt_identity()
    assessment = MadyaAssessment.query.filter_by(id=assessment_id, user_id=current_user_id).first_or_404("Asesmen Madya tidak ditemukan atau bukan milik Anda.")

    risk_entries = RiskInputMadya.query.filter_by(assessment_id=assessment.id).order_by(RiskInputMadya.id).all()

    result = [format_risk_input_entry(entry) for entry in risk_entries]
    # for entry in risk_entries:
    #     entry_dict = {c.name: getattr(entry, c.name) for c in entry.__table__.columns}
    #     for date_field in ['tanggal_identifikasi', 'jadwal_mulai_penanganan', 'jadwal_selesai_penanganan', 'tanggal_review']:
    #         if entry_dict.get(date_field):
    #             entry_dict[date_field] = entry_dict[date_field].isoformat()
    #     result.append(entry_dict)

    return jsonify(result)

@risk_management_levels_bp.route('/risk-inputs/<int:risk_input_id>', methods=['PUT'])
@jwt_required()
def update_risk_input(risk_input_id):
    """Memperbarui satu entri Risk Input."""
    current_user_id = get_jwt_identity()
    risk_input = RiskInputMadya.query.get_or_404(risk_input_id)

    assessment = MadyaAssessment.query.get_or_404(risk_input.assessment_id)
    if str(assessment.user_id) != current_user_id:
        return jsonify({"msg": "Akses ditolak."}), 403

    data = request.get_json()
    if not data:
        return jsonify({"msg": "Data tidak boleh kosong."}), 400

    # calculated_values = calculate_risk_values(data, assessment.risk_map_template_id)

# 1. Parse semua input numerik yang relevan dari data BARU atau data LAMA
    inherent_p = parse_int_or_none(data.get('inherent_probabilitas', risk_input.inherent_probabilitas))
    inherent_i = parse_int_or_none(data.get('inherent_dampak', risk_input.inherent_dampak))
    inherent_prob_kualitatif = parse_float_or_none(data.get('inherent_prob_kualitatif', risk_input.inherent_prob_kualitatif))
    inherent_dampak_finansial = parse_float_or_none(data.get('inherent_dampak_finansial', risk_input.inherent_dampak_finansial))
    residual_p = parse_int_or_none(data.get('residual_probabilitas', risk_input.residual_probabilitas))
    residual_i = parse_int_or_none(data.get('residual_dampak', risk_input.residual_dampak))
    residual_prob_kualitatif = parse_float_or_none(data.get('residual_prob_kualitatif', risk_input.residual_prob_kualitatif))
    biaya_penanganan = parse_float_or_none(data.get('biaya_penanganan', risk_input.biaya_penanganan))

    # 2. Hitung Skor menggunakan template (gunakan data yang sudah diparsing)
    score_data_for_calc = {
        'inherent_probabilitas': inherent_p, 'inherent_dampak': inherent_i,
        'residual_probabilitas': residual_p, 'residual_dampak': residual_i
    }
    calculated_scores = calculate_scores_from_template(score_data_for_calc, assessment.risk_map_template_id)
    inherent_skor = calculated_scores.get('inherent_skor')
    residual_skor = calculated_scores.get('residual_skor')

    # 3. Hitung Nilai Bersih Inheren
    inherent_nilai_bersih = None
    if inherent_prob_kualitatif is not None and inherent_dampak_finansial is not None:
        inherent_nilai_bersih = inherent_dampak_finansial * (inherent_prob_kualitatif / 100.0)

    # 4. Hitung Dampak Finansial Residual (sesuai logika frontend)
    calculated_residual_dampak_finansial = None
    if residual_prob_kualitatif is not None and inherent_dampak_finansial is not None:
        calculated_residual_dampak_finansial = inherent_dampak_finansial * (residual_prob_kualitatif / 100.0)

    # 5. Hitung Nilai Bersih Residual (sesuai logika frontend)
    residual_nilai_bersih = None
    if residual_prob_kualitatif is not None and calculated_residual_dampak_finansial is not None:
        residual_nilai_bersih = calculated_residual_dampak_finansial * (residual_prob_kualitatif / 100.0)
            
    # Update fields based on request data
    risk_input.sasaran_id = parse_int_or_none(data.get('sasaran_id', risk_input.sasaran_id))
    risk_input.kode_risiko = data.get('kode_risiko', risk_input.kode_risiko)
    risk_input.status_risiko = data.get('status_risiko', risk_input.status_risiko)
    risk_input.peluang_ancaman = data.get('peluang_ancaman', risk_input.peluang_ancaman)
    risk_input.kategori_risiko = data.get('kategori_risiko', risk_input.kategori_risiko)
    risk_input.kategori_risiko_lainnya = data.get('kategori_risiko_lainnya') if risk_input.kategori_risiko == 'Lainnya' else None
    risk_input.unit_kerja = data.get('unit_kerja', risk_input.unit_kerja)
    risk_input.tanggal_identifikasi = parse_date_or_none(data.get('tanggal_identifikasi')) or risk_input.tanggal_identifikasi
    risk_input.deskripsi_risiko = data.get('deskripsi_risiko', risk_input.deskripsi_risiko)
    risk_input.akar_penyebab = data.get('akar_penyebab', risk_input.akar_penyebab)
    risk_input.indikator_risiko = data.get('indikator_risiko', risk_input.indikator_risiko)
    risk_input.internal_control = data.get('internal_control', risk_input.internal_control)
    risk_input.deskripsi_dampak = data.get('deskripsi_dampak', risk_input.deskripsi_dampak)
    risk_input.pemilik_risiko = data.get('pemilik_risiko', risk_input.pemilik_risiko)
    risk_input.jabatan_pemilik = data.get('jabatan_pemilik', risk_input.jabatan_pemilik)
    risk_input.kontak_pemilik_hp = data.get('kontak_pemilik_hp', risk_input.kontak_pemilik_hp)
    risk_input.kontak_pemilik_email = data.get('kontak_pemilik_email', risk_input.kontak_pemilik_email)
    risk_input.strategi = data.get('strategi', risk_input.strategi)
    risk_input.rencana_penanganan = data.get('rencana_penanganan', risk_input.rencana_penanganan)
    risk_input.penanganan_dilakukan = data.get('penanganan_dilakukan', risk_input.penanganan_dilakukan)
    risk_input.status_penanganan = data.get('status_penanganan', risk_input.status_penanganan)
    risk_input.jadwal_mulai_penanganan = parse_date_or_none(data.get('jadwal_mulai_penanganan')) or risk_input.jadwal_mulai_penanganan
    risk_input.jadwal_selesai_penanganan = parse_date_or_none(data.get('jadwal_selesai_penanganan')) or risk_input.jadwal_selesai_penanganan
    risk_input.pic_penanganan = data.get('pic_penanganan', risk_input.pic_penanganan)
    risk_input.tanggal_review = parse_date_or_none(data.get('tanggal_review')) or risk_input.tanggal_review
    
    risk_input.inherent_probabilitas = inherent_p
    risk_input.inherent_dampak = inherent_i
    risk_input.inherent_skor = inherent_skor
    risk_input.inherent_prob_kualitatif = inherent_prob_kualitatif
    risk_input.inherent_dampak_finansial = inherent_dampak_finansial
    risk_input.inherent_nilai_bersih = inherent_nilai_bersih
    risk_input.biaya_penanganan = biaya_penanganan
    risk_input.residual_probabilitas = residual_p
    risk_input.residual_dampak = residual_i
    risk_input.residual_skor = residual_skor
    risk_input.residual_prob_kualitatif = residual_prob_kualitatif
    risk_input.residual_dampak_finansial = calculated_residual_dampak_finansial # Update dengan hasil hitung
    risk_input.residual_nilai_bersih = residual_nilai_bersih
    risk_input.tanggal_review = parse_date_or_none(data.get('tanggal_review')) or risk_input.tanggal_review

    db.session.commit()
    db.session.refresh(risk_input)
    
    update_sasaran_kpi_scores(risk_input.sasaran_id)
    db.session.commit()

    # Kembalikan data yang sudah diupdate
    # updated_entry_dict = {c.name: getattr(risk_input, c.name) for c in risk_input.__table__.columns}
    # for date_field in ['tanggal_identifikasi', 'jadwal_mulai_penanganan', 'jadwal_selesai_penanganan', 'tanggal_review']:
    #     if updated_entry_dict.get(date_field):
    #         updated_entry_dict[date_field] = updated_entry_dict[date_field].isoformat()

    return jsonify({"msg": "Risk Input berhasil diperbarui.", "entry": format_risk_input_entry(risk_input)})


@risk_management_levels_bp.route('/risk-inputs/<int:risk_input_id>', methods=['DELETE'])
@jwt_required()
def delete_risk_input(risk_input_id):
    """Menghapus satu entri Risk Input."""
    current_user_id = get_jwt_identity()
    risk_input = RiskInputMadya.query.get_or_404(risk_input_id)

    assessment = MadyaAssessment.query.get_or_404(risk_input.assessment_id)
    if str(assessment.user_id) != current_user_id:
        return jsonify({"msg": "Akses ditolak."}), 403

    db.session.delete(risk_input)
    db.session.commit()
    return jsonify({"msg": "Risk Input berhasil dihapus."})